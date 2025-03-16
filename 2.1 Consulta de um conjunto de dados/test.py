import sqlite3
import pandas as pd
import os
from datetime import datetime, timedelta 
import plotly.express as px 
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image

db_path = "database/retrun.db"
excel_path = "database/retrun.xlsx"




if not os.path.exists(db_path):
    df = pd.read_excel(excel_path, sheet_name="Retorno SVC São Carlos ")
    df = df[["DATA (OBRIGATORIO)", "ID", "MOTIVO","Devolvido", "Status Deixado ", "Contador", "Status", "Rota",  "Transportadora"]]
    cols = ["MOTIVO", "Devolvido", "Status Deixado ", "Status", "Rota", "Transportadora"]

    df[cols] = df[cols].apply(lambda col: col.str.strip().str.upper() if col.dtype == 'object' else col)
    df.loc[df["Rota"] == "-", "Rota"] = "ROTA BRANCA"
    df = df.dropna()
    
    df["MOTIVO"] = df["MOTIVO"].apply(lambda x: x.split("|")[0])

    conn = sqlite3.connect(db_path)
    c = conn.cursor()

    df.to_sql("returns", conn, if_exists="replace", index=False)

    conn.close()



def Mot_analyzes():
    conn = sqlite3.connect(db_path)
    c = conn.cursor()

    command = f'''
    SELECT MOTIVO, COUNT(*) as cat_count
    FROM returns       
    GROUP BY MOTIVO
    HAVING cat_count > 10
    ORDER BY cat_count DESC 
    '''

    c.execute(command)

    rows = c.fetchall()

    motivos = [row[0] for row in rows]
    contagens = [row[1] for row in rows]              
       

    newDate = datetime.today()
    newDate = newDate.strftime("%Y-%m-%d")
    

    output_file = f'graphics motivos {newDate}.xlsx'

    try:
        workbook = openpyxl.load_workbook(output_file)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    
    worksheet = workbook.active
    worksheet.title = "Report"

    worksheet.append(["Motivo", "Quantidade"])
    for motivo, contagem in zip(motivos, contagens):
        worksheet.append([motivo, contagem])
    try: 
        fig = px.funnel(x=motivos, y=contagens,  color=contagens,
                    labels={'x': 'Motivo', 'y': 'Contagem'},
                    title=f'Motivos de Retorno para a Data {newDate}',
                    text=contagens
                    )
    
        image_path = f'graphics/images/report image- {newDate}.png'
        # Atualizar o layout do gráfico
        fig.update_traces(texttemplate='%{text}', textposition='outside')
        fig.update_layout(uniformtext_minsize=8, uniformtext_mode='hide')
        fig.write_html(f'graphics/graphic/Report graphic - {newDate}.html')       
        fig.write_image(image_path, width=1200, height=700)       
       
        img = Image(image_path)
        worksheet.add_image(img, 'D1')  # Inserir a imagem na célula A1
     
        workbook.save(output_file)
        print(f"Arquivo Excel '{output_file}' criado com sucesso!")     
    except ValueError as e:
        print("Empy date!")
        print(e)
    
    conn.close()
    
    #fig.show()

def Transanalyzes():
    conn = sqlite3.connect(db_path)
    c = conn.cursor()

    command = f'''
    SELECT Transportadora, COUNT(*) as cat_count
    FROM returns       
    GROUP BY Transportadora    
    ORDER BY cat_count DESC 
    '''

    c.execute(command)

    rows = c.fetchall()

    motivos = [row[0] for row in rows]
    contagens = [row[1] for row in rows]              
       

    newDate = datetime.today()
    newDate = newDate.strftime("%Y-%m-%d")
    

    output_file = f'graphics motivos {newDate}.xlsx'

    try:
        workbook = openpyxl.load_workbook(output_file)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    
    worksheet = workbook.active
    worksheet.title = "Report"

    worksheet.append(["Motivo", "Quantidade"])
    for motivo, contagem in zip(motivos, contagens):
        worksheet.append([motivo, contagem])
    try: 
        fig = px.funnel(x=motivos, y=contagens,  color=contagens,
                    labels={'x': 'Motivo', 'y': 'Contagem'},
                    title=f'Motivos de Retorno para a Data {newDate}',
                    text=contagens
                    )
    
        image_path = f'graphics/images/report image- {newDate}.png'
        # Atualizar o layout do gráfico
        fig.update_traces(texttemplate='%{text}', textposition='outside')
        fig.update_layout(uniformtext_minsize=8, uniformtext_mode='hide')
        fig.write_html(f'graphics/graphic/Report graphic - {newDate}.html')       
        fig.write_image(image_path, width=1200, height=700)       
       
        img = Image(image_path)
        worksheet.add_image(img, 'D1')  # Inserir a imagem na célula A1
     
        workbook.save(output_file)
        print(f"Arquivo Excel '{output_file}' criado com sucesso!")     
    except ValueError as e:
        print("Empy date!")
        print(e)
    
    conn.close()
    
    #fig.show()



#Mot_analyzes()
Transanalyzes()
